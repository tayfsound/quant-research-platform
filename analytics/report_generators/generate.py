"""Rapor üreticiler."""
import csv
import json


class ReportGenerator:
    @staticmethod
    def to_json(metrics: dict, path: str):
        with open(path, "w") as f:
            json.dump(metrics, f, indent=2)

    @staticmethod
    def to_csv(fills: list[dict], path: str):
        if not fills:
            return
        with open(path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fills[0].keys())
            writer.writeheader()
            writer.writerows(fills)

    @staticmethod
    def to_excel(metrics: dict, fills: list[dict], path: str):
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Metrics"
        for i, (k, v) in enumerate(metrics.items(), 1):
            ws1.cell(row=i, column=1, value=k)
            ws1.cell(row=i, column=2, value=v)
        ws2 = wb.create_sheet("Fills")
        if fills:
            for j, key in enumerate(fills[0].keys(), 1):
                ws2.cell(row=1, column=j, value=key)
            for i, f in enumerate(fills, 2):
                for j, (k, v) in enumerate(f.items(), 1):
                    ws2.cell(row=i, column=j, value=v)
        wb.save(path)
